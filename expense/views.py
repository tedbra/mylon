# expense/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Sum, F, DecimalField
from .models import Requisition, RequisitionItem, ExpenseCategory
from .forms import RequisitionForm, RequisitionItemFormSet
from django.contrib import messages
from students.models import Campus

#For exporting data to Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone # 🎯 Added for dynamic filename dating

# expense/views.py
@login_required
def requisition_create_view(request):
    user_profile = getattr(request.user, 'profile', None)
    is_director = user_profile.is_director if user_profile else False
    is_global_admin = user_profile.is_global_admin if user_profile else False

    if request.method == 'POST':
        # Safely reset formset extras to accept the exact post count payload
        RequisitionItemFormSet.extra = 0
        
        form = RequisitionForm(request.POST, user=request.user)
        formset = RequisitionItemFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    # 1. Instantiate the requisition parent without committing to DB yet
                    requisition = form.save(commit=False)
                    
                    # 2. Assign the Requester / Profile Owner 
                    requisition.requester = request.user
                    
                    # 3. Guardrail: If they aren't a Director/Founder, force their profile campus branch
                    if not is_director:
                        requisition.campus = user_profile.campus
                    
                    # 4. Explicitly set initial status to PENDING
                    requisition.status = 'PENDING'
                    
                    # 5. Save the parent to generate its database Primary Key ID
                    requisition.save()
                    
                    # 6. Bind the newly saved parent record to the inline formset instances and save them
                    formset.instance = requisition
                    formset.save()
                    
                messages.success(request, "Weekly operational requisition submitted successfully.")
                
                # Redirect based on authority tier mapping back to their perspective boards
                if is_director or is_global_admin:
                    return redirect('expense_report_admin')
                return redirect('expense_report')
                
            except Exception as e:
                messages.error(request, f"Database Transaction Error: {str(e)}")
        else:
            # Extract errors clearly if submission misbehaves
            messages.error(request, "Submission failed! Please review row fields or structural settings below.")
            for field, errors in form.errors.items():
                messages.error(request, f"Header field [{field}]: {', '.join(errors)}")
            for i, form_errors in enumerate(formset.errors):
                for field, errors in form_errors.items():
                    messages.error(request, f"Row #{i+1} [{field}]: {', '.join(errors)}")
    else:
        # GET request handling initialization state
        form = RequisitionForm(user=request.user)
        
        # Inject exactly one fresh empty row to welcome input configuration layouts
        RequisitionItemFormSet.extra = 1
        formset = RequisitionItemFormSet()

    context = {
        'form': form,
        'formset': formset,
        'title': 'Apply Weekly Operational Requisition'
    }
    return render(request, 'expense/requisition_form.html', context)


@login_required
def update_requisition_status(request, pk, status_choice):
    """
    Executive authorization switch action for head office directors.
    """
    if not request.user.profile.is_director:
        return redirect('expense_report')
        
    requisition = get_object_or_404(Requisition.global_objects, pk=pk)
    if status_choice in ['APPROVED', 'REJECTED']:
        requisition.status = status_choice
        requisition.save()
    return redirect('expense_report')


@login_required
def xrequisition_update_view(request, pk):
    """
    Allows the creator or admin to update a requisition and its line items,
    provided the status is still 'PENDING'.
    """
    user_profile = getattr(request.user, 'profile', None)
    is_global = user_profile.is_director if user_profile else getattr(request.user, 'is_director', False)
    
    # 1. Fetch the record based on user clearance level
    if is_global:
        requisition = get_object_or_404(Requisition.global_objects, pk=pk)
    else:
        requisition = get_object_or_404(Requisition.objects, pk=pk)

    # 2. CRITICAL BUSINESS PROTECTION RULE: Lock down approved/rejected records
    if requisition.status != 'PENDING':
        messages.error(request, "Security Lock: This requisition has already been processed and cannot be modified.")
        return redirect('expense_report')

    if request.method == 'POST':
        # 🎯 FIX: Pass the user keyword argument to match the updated RequisitionForm constructor!
        form = RequisitionForm(request.POST, instance=requisition, user=request.user)
        formset = RequisitionItemFormSet(request.POST, instance=requisition)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    updated_requisition = form.save(commit=False)
                    updated_requisition.save()
                    formset.save()
                messages.success(request, "Requisition updated successfully.")
                if is_global:
                    return redirect('expense_report_admin')
                else:
                    return redirect('expense_report')
            except Exception as e:
                messages.error(request, f"Database Update Error: {str(e)}")
        else:
            # 💡 Diagnostic Tool: If validation fails, dump errors clearly as messages
            # so you can see exactly which line field is breaking submission.
            for field, errors in form.errors.items():
                messages.error(request, f"Form field [{field}]: {', '.join(errors)}")
            
            for i, form_errors in enumerate(formset.errors):
                for field, errors in form_errors.items():
                    messages.error(request, f"Item Row #{i+1} field [{field}]: {', '.join(errors)}")
    else:
        # 🎯 FIX: Pass the user keyword argument on initial GET instantiation states as well!
        form = RequisitionForm(instance=requisition, user=request.user)
        formset = RequisitionItemFormSet(instance=requisition)

    context = {
        'form': form,
        'formset': formset,
        'requisition': requisition,
        'title': f'Modify Requisition Sheet #00{requisition.id}'
    }
    return render(request, 'expense/requisition_update.html', context)


@login_required
def expense_report_view(request):
    """
    Acts as a secure traffic controller. Redirects global administrators
    to the executive campus-grouped dashboard, while keeping local campus staff 
    on the single-tenant regional ledger view.
    """
    user_profile = getattr(request.user, 'profile', None)
    is_global = user_profile.is_director if user_profile else getattr(request.user, 'is_director', False)

    if is_global:
        return redirect('expense_report_admin')

    # --- LOCAL REGIONAL STAFF VIEW LOGIC ---
    if user_profile and user_profile.campus:
        base_queryset = Requisition.objects.filter(campus=user_profile.campus)
    elif hasattr(request.user, 'campus') and request.user.campus:
        base_queryset = Requisition.objects.filter(campus=request.user.campus)
    else:
        base_queryset = Requisition.objects.none()

    requisitions = base_queryset.select_related('campus', 'requester')
    total_requested_funds = sum(req.total_cost for req in requisitions)
    approved_funds = sum(req.total_cost for req in requisitions.filter(status='APPROVED'))
    pending_count = requisitions.filter(status='PENDING').count()

    context = {
        'requisitions': requisitions,
        'total_requested_funds': total_requested_funds,
        'approved_funds': approved_funds,
        'pending_count': pending_count,
        'is_director': False,
        'user_campus': user_profile.campus if user_profile else None
    }
    return render(request, 'expense/expense_report.html', context)


@login_required
def expense_report_admin_view(request):
    """
    Executive Multi-Tenant Boardroom View. Grouping costs and voucher rows
    individually by campus for high-level cross-examination.
    """
    user_profile = getattr(request.user, 'profile', None)
    is_global = user_profile.is_director if user_profile else getattr(request.user, 'is_director', False)
    is_global_admin = user_profile.is_global_admin if user_profile else getattr(request.user, 'is_global_admin', False)
    
    # Security constraint: Kick out anyone who isn't an admin
    if not is_global:
        return redirect('expense_report')

    all_campuses = Campus.objects.all()
    campus_blocks = []
    
    # Global metrics across the entire organization
    grand_total_requested = 0
    grand_total_approved = 0
    grand_total_pending_vouchers = 0

    for campus in all_campuses:
        # Pull all vouchers matching this specific campus block
        campus_requisitions = Requisition.global_objects.filter(campus=campus).select_related('requester')
        
        # Calculate individual localized metrics for this block
        campus_requested = sum(req.total_cost for req in campus_requisitions)
        campus_approved = sum(req.total_cost for req in campus_requisitions.filter(status='APPROVED'))
        campus_pending_count = campus_requisitions.filter(status='PENDING').count()

        # Append structured map data packet
        campus_blocks.append({
            'campus': campus,
            'requisitions': campus_requisitions,
            'total_requested': campus_requested,
            'total_approved': campus_approved,
            'pending_count': campus_pending_count
        })

        # Aggregate grand organizational sums
        grand_total_requested += campus_requested
        grand_total_approved += campus_approved
        grand_total_pending_vouchers += campus_pending_count

    context = {
        'campus_blocks': campus_blocks,
        'is_global_admin': is_global_admin,
        'grand_total_requested': grand_total_requested,
        'grand_total_approved': grand_total_approved,
        'grand_total_pending_vouchers': grand_total_pending_vouchers
    }
    return render(request, 'expense/expense_report_admin.html', context)


@login_required
def requisition_update_view(request, pk):
    """
    Allows the creator or admin to update a requisition and its line items,
    provided the status is still 'PENDING'.
    """
    user_profile = getattr(request.user, 'profile', None)
    is_global = user_profile.is_director if user_profile else getattr(request.user, 'is_director', False)
    
    if is_global:
        requisition = get_object_or_404(Requisition.global_objects, pk=pk)
    else:
        get_object_or_404(Requisition.objects, pk=pk)

    if requisition.status != 'PENDING':
        messages.error(request, "Security Lock: This requisition has already been processed and cannot be modified.")
        if is_global:
            return redirect('expense_report_admin')
        return redirect('expense_report')

    # Force extra count to zero for updates
    RequisitionItemFormSet.extra = 0

    if request.method == 'POST':
        form = RequisitionForm(request.POST, instance=requisition, user=request.user)
        
        # Admin vs Local user formset initialization
        if is_global:
            from .models import RequisitionItem
            qs = RequisitionItem.global_objects.filter(requisition=requisition) if hasattr(RequisitionItem, 'global_objects') else RequisitionItem.objects.filter(requisition=requisition)
            formset = RequisitionItemFormSet(request.POST, instance=requisition, queryset=qs)
        else:
            formset = RequisitionItemFormSet(request.POST, instance=requisition)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    updated_requisition = form.save(commit=False)
                    updated_requisition.save()
                    formset.save()
                messages.success(request, "Requisition updated successfully.")
                if is_global:
                    return redirect('expense_report_admin')
                return redirect('expense_report')
            except Exception as e:
                messages.error(request, f"Database Update Error: {str(e)}")
        else:
            # 🎯 AGGRESIVE VISUAL ERROR CATCHER
            messages.error(request, "❌ SUBMISSION REJECTED BY DJANGO VALIDATION ENGINE:")
            
            # 1. Capture Parent Form Errors
            for field, errors in form.errors.items():
                messages.error(request, f"➔ Parent Field [{field}]: {', '.join(errors)}")
            if form.non_field_errors():
                messages.error(request, f"➔ Parent Global Error: {', '.join(form.non_field_errors())}")
            
            # 2. Capture Child Inline Row Errors
            for i, form_errors in enumerate(formset.errors):
                if form_errors:
                    for field, errors in form_errors.items():
                        messages.error(request, f"➔ Row #{i+1} Field [{field}]: {', '.join(errors)}")
            
            # 3. Capture Structural Formset Errors (Management Form / Security Mismatches)
            if formset.non_form_errors():
                messages.error(request, f"➔ Formset Structure Error: {', '.join(formset.non_form_errors())}")
                
    else:
        form = RequisitionForm(instance=requisition, user=request.user)
        if is_global:
            from .models import RequisitionItem
            qs = RequisitionItem.global_objects.filter(requisition=requisition) if hasattr(RequisitionItem, 'global_objects') else RequisitionItem.objects.filter(requisition=requisition)
            formset = RequisitionItemFormSet(instance=requisition, queryset=qs)
        else:
            formset = RequisitionItemFormSet(instance=requisition)

    context = {
        'form': form,
        'formset': formset,
        'requisition': requisition,
        'title': f'Modify Requisition Sheet #00{requisition.id}'
    }
    return render(request, 'expense/requisition_update.html', context)


@login_required
def requisition_detail_view(request, pk):
    """
    Read-only presentation mode for checking a requisition voucher 
    and approving/rejecting/editing it via top panel macro-controls.
    """
    user_profile = getattr(request.user, 'profile', None)
    is_director = user_profile.is_director if user_profile else getattr(request.user, 'is_director', False)
    is_global_admin = user_profile.is_global_admin if user_profile else getattr(request.user, 'is_global_admin', False)
    
    # Fetch based on clearance
    if is_director:
        requisition = get_object_or_404(Requisition.global_objects, pk=pk)
    else:
        requisition = get_object_or_404(Requisition.objects, pk=pk)
        
    # Prefetch child items for optimized page loading speed
    items = requisition.items.all() 
    
    # Calculate costs inline per item to pass to the template context explicitly
    item_details = []
    grand_total = 0
    for item in items:
        line_total = item.quantity * item.price_per_unit
        grand_total += line_total
        item_details.append({
            'instance': item,
            'line_total': line_total
        })

    context = {
        'requisition': requisition,
        'item_details': item_details,
        'grand_total': grand_total,
        'is_director': is_director,
        'is_global_admin': is_global_admin,
        'title': f'Review Requisition Sheet #00{requisition.id}'
    }
    return render(request, 'expense/requisition_detail.html', context)



@login_required
def export_requisitions_excel(request):
    """
    Exports procurement requisitions grouped into separate worksheet tabs 
    by campus, utilizing a denormalized flat-table structure per sheet.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Resolve multi-campus permission visibility scopes
    user_profile = request.user.profile
    if user_profile.is_director or user_profile.is_global_admin:
        campuses = Campus.objects.all()
    else:
        campuses = Campus.objects.filter(id=user_profile.campus.id) if user_profile.campus else Campus.objects.none()

    # Styling Profiles
    purple_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid") # Deep Purple Header
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F9F5FF", end_color="F9F5FF", fill_type="solid") # Soft Purple Zebra Striping
    normal_font = Font(name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
                         top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    headers = [
        "Requisition ID", "Operational Week", "Date Created", "Applicant", 
        "Approval Status", "Expense Item", "Category", "Quantity", 
        "Price Per Unit", "Total Line Cost"
    ]

    # 2. Iterate through Campuses to build isolated worksheet tabs
    for campus in campuses:
        # Clean up tab name to fit Excel rules (max 31 chars, no special characters)
        sheet_title = "".join(c for c in campus.name if c.isalnum() or c in "._- ")[:31].strip() or f"Campus {campus.id}"

        # Fetch all vouchers belonging strictly to this loop's campus branch
        requisitions = Requisition.global_objects.filter(campus=campus).select_related('term').prefetch_related('items__category')

        # Skip generating an empty tab if this specific campus branch has zero requisitions log rows
        if not requisitions.exists():
            continue

        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        ws.append(headers)

        # Apply header formatting
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = white_bold; cell.fill = purple_fill; cell.border = thin_border

        row_counter = 2
        is_even_voucher = False

        # 3. Populate rows for this campus sheet tab
        for req in requisitions:
            line_items = req.items.all()

            week_str = req.calendar_week.strftime('%Y-%m-%d') if req.calendar_week else "N/A"
            created_str = req.created_at.strftime('%Y-%m-%d') if req.created_at else "N/A"
            applicant_str = req.requester_historical_name or "Unknown"
            status_str = req.get_status_display()

            if not line_items.exists():
                row_data = [f"REQ-{req.id:04d}", week_str, created_str, applicant_str, status_str, "No Items Specified", "N/A", 0, 0.00, 0.00]
                ws.append(row_data)
                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_counter, column=col_idx)
                    cell.font = normal_font; cell.border = thin_border
                    if is_even_voucher: cell.fill = zebra_fill
                row_counter += 1
                continue

            for item in line_items:
                qty = int(item.quantity or 0)
                unit_price = float(item.price_per_unit or 0)
                line_total = float(item.cost or 0)

                row_data = [
                    f"REQ-{req.id:04d}",
                    week_str,
                    created_str,
                    applicant_str,
                    status_str,
                    item.expense_name,
                    item.category.title if item.category else "N/A",  # 🎯 FIXED: Correctly using .title now
                    qty,
                    unit_price,
                    line_total
                ]
                ws.append(row_data)

                for col_idx in range(1, len(row_data) + 1):
                    cell = ws.cell(row=row_counter, column=col_idx)
                    cell.font = normal_font; cell.border = thin_border
                    if is_even_voucher:
                        cell.fill = zebra_fill

                    # Right align numeric metrics (columns 8, 9, 10 in this layout version)
                    if col_idx in [8, 9, 10]:
                        cell.alignment = Alignment(horizontal="right")
                        if col_idx in [9, 10]:
                            cell.number_format = '#,##0.00'

                row_counter += 1

            # Alternate background tint tracking variables per full requisition block group
            is_even_voucher = not is_even_voucher

        # Auto-adjust column widths dynamically for this specific worksheet tab instance
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # 4. Fallback safe check to handle empty environments smoothly
    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Requisitions Logged")
        ws.append(["No matching requisition entries exist within your authorization boundaries."])

    # 5. Build dynamic file attachment layout names
    current_date = timezone.now().strftime('%Y-%m-%d')
    filename = f"Campus_Requisitions_Log_{current_date}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from django.shortcuts import render
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.db.models import Sum, Count, Q,F, ExpressionWrapper, DecimalField
from .models import Student, Campus, Term, Invoice, PaymentHistory
from expense.models import Requisition, RequisitionItem
from django.contrib import messages
from django.shortcuts import render, redirect


@login_required    
def global_analytics_dashboard(request):
    """
    Computes school-wide term operational analytics, matching collected recurring revenue 
    metrics and term-based expenditure breakdowns stratified by actual term footprints.
    """
    user_profile = getattr(request.user, 'profile', None)
    is_director = user_profile.is_director if user_profile else False
    is_global_admin = user_profile.is_global_admin if user_profile else False

    if not is_director and not is_global_admin:
        messages.error(request, "Access Denied: Analytics dashboards are restricted to Directors and Founders.")
        return redirect('revenue_dashboard')

    # 1. 🎯 ANCHOR BY ACTIVE INVOICE FOOTPRINTS (Captures returning + new students dynamically per term)
    active_term_footprints = Invoice.objects.values(
        'student__campus__id',
        'student__campus__name', 
        'term__session__name', 
        'term__term_name',
        'term__id'
    ).annotate(
        total_active_students=Count('student', distinct=True)
    ).order_by('student__campus__id', '-term__session__name', '-term__term_name')

    # 2. 💰 AGGREGATE REVENUE DYNAMICALLY
    # Since PaymentHistory lacks an invoice FK, we pull payments and group them by campus and payment year
    revenue_data = PaymentHistory.global_objects.values(
        'student__campus__name',
        'date_paid__year'
    ).annotate(
        total_revenue=Sum('amount_paying')
    )

    # Build a revenue lookup dictionary based on Campus and Year
    revenue_lookup = {}
    for row in revenue_data:
        campus = row['student__campus__name'] or "Unassigned Campus"
        pay_year = row['date_paid__year']
        
        lookup_key = (campus, pay_year)
        revenue_lookup[lookup_key] = revenue_lookup.get(lookup_key, Decimal('0.00')) + (row['total_revenue'] or Decimal('0.00'))

    # 3. 🏢 AGGREGATE OPERATIONAL EXPENSES PER CAMPUS/TERM
    expense_lookup = {}
    approved_requisitions = Requisition.global_objects.filter(
        status='APPROVED'
    ).select_related('campus', 'term__session')
    
    for req in approved_requisitions:
        c_name = req.campus.name if req.campus else "Unassigned Campus"
        session_name = req.term.session.name if req.term and req.term.session else "Unknown Session"
        term_name = req.term.term_name if req.term else "Unknown Term"
        
        expense_key = (c_name, session_name, term_name)
        expense_lookup[expense_key] = expense_lookup.get(expense_key, Decimal('0.00')) + Decimal(str(req.total_cost))

    # 4. MAP DATA STRUCTURALLY FOR THE TEMPLATE RENDER LOOP
    unordered_campus_sections = {}
    campus_pk_lookup = {}
    for row in active_term_footprints:
        campus_name = row['student__campus__name'] or "Unassigned Campus"
        campus_id = row['student__campus__id']
        session_name = row['term__session__name'] or "Unknown Session"
        term_name = row['term__term_name'] or "Unknown Term"
        term_id = row['term__id']

        # Map tracking PK to build your campus sorted sequence safely
        campus_pk_lookup[campus_name] = campus_id
        
        # Parse the calendar year out of your session name string (e.g., "2026/2027" -> 2026)
        try:
            calendar_year = int(session_name.split('/')[0])
        except (ValueError, IndexError):
            calendar_year = 2026

        # Match lookups safely
        rev_key = (campus_name, calendar_year)
        exp_key = (campus_name, session_name, term_name)
        
        # 🎯 Since payments are yearly/date-based here, we attribute revenue to the active session year row
        revenue_collected = Decimal(str(revenue_lookup.get(rev_key, 0.00))) if 'TERM_1' in term_name else Decimal('0.00')
        # Note: If your template splits revenue evenly or you track specific terms, you can adjust the line above!
        
        operational_expenses = Decimal(str(expense_lookup.get(exp_key, 0.00)))
        
        processed_row = {
            'campus_id': campus_id,
            'term_id': term_id,
            'enrollment_year': session_name, 
            'enrollment_term': term_name,
            'total_intake': row['total_active_students'], 
            'revenue_collected': revenue_collected if 'TERM_1' in term_name else Decimal(str(revenue_lookup.get(rev_key, 0.00))) / 3, # Fallback: split revenue across terms evenly if collected globally
            'operational_expenses': operational_expenses,
            'net_operating_margin': revenue_collected - operational_expenses,
            'the_year': calendar_year,
        }
        
        if campus_name not in unordered_campus_sections:
            unordered_campus_sections[campus_name] = []
        unordered_campus_sections[campus_name].append(processed_row)

    campus_sections = {
        campus_name: unordered_campus_sections[campus_name]
        for campus_name in sorted(
            unordered_campus_sections.keys(), 
            key=lambda name: campus_pk_lookup.get(name, 0)
        )
    }

    context = {
        'campus_sections': campus_sections,
        'title': 'Global Analytics Matrix & Operating P&L Ledger'
    }
    return render(request, 'students/analytics_report.html', context)



def get_term_summary_data(campus_id, term_id):
    """
    Helper function to abstract the heavy aggregate database mathematical formulas
    so both the HTML view and Excel exporter can share the exact same data matrix.
    """
    campus = Campus.objects.get(id=campus_id)
    term = Term.objects.get(id=term_id)

    # ==========================================
    # SECTION 1: REVENUE AND LEDGER AGGREGATION
    # ==========================================
    students_query = Student.objects.filter(campus=campus)
    
    # Run a unified database aggregate loop mapping directly to your invoice properties
    student_metrics = students_query.aggregate(
        total_pupils=Count('id'),
        active_pupils = Count('invoices__id', filter=Q(invoices__term=term, status='ACTIVE')),
        #active_pupils=Count('id', filter=Q(status='ACTIVE')),
        on_transport=Count('invoices__id', filter=Q(invoices__term=term, invoices__transport_amount__gt=0)),
        approved_scholarships_count=Count('invoices__id', filter=Q(invoices__term=term, invoices__scholarship_applied__gt=0)),
        total_arrears=Sum('invoices__previous_arrears', filter=Q(invoices__term=term)),
        total_scholarships_given=Sum('invoices__scholarship_applied', filter=Q(invoices__term=term)),
        total_tuition=Sum('invoices__tuition_amount', filter=Q(invoices__term=term)),
        total_fees_billed=Sum('invoices__amount', filter=Q(invoices__term=term)), # Grand Total Invoice amount
    )

    # Fallback to zero values if no students or invoices exist yet
    active_count = student_metrics['active_pupils'] or 0
    transport_count = student_metrics['on_transport'] or 0
    scholarship_pupils_count = student_metrics['approved_scholarships_count'] or 0
    
    arrears_sum = student_metrics['total_arrears'] or Decimal('0.00')
    scholarship_amt_sum = student_metrics['total_scholarships_given'] or Decimal('0.00')
    total_due_sum = student_metrics['total_fees_billed'] or Decimal('0.00')

    # 🎯 1. Pure Gross Term Billings (Gross Tuition + Transport + Extras before deductions)
    pure_term_billings = total_due_sum - arrears_sum + scholarship_amt_sum

    # 🎯 2. Aggregate total transport metrics billed via invoice components
    transport_sum = students_query.filter(invoices__term=term).aggregate(
        total_trans=Sum('invoices__transport_amount')
    )['total_trans'] or Decimal('0.00')

    # 🎯 3. Calculate Base Term Tuition Fees securely
    base_term_fees = student_metrics['total_tuition'] or Decimal('0.00')

    # 🎯 4. FIX: Extras sum is the true remainder when stripping known Tuition and Transport from gross billings
    extras_sum = max(Decimal('0.00'), pure_term_billings - transport_sum - base_term_fees)

    # Calculate real payment logs captured within this term lifecycle window boundaries
    collected_sum = Decimal('0.00')
    if term.start_date and term.end_date:
        collected_sum = Student.global_objects.filter(campus=campus).filter(
            payments__date_paid__range=(term.start_date, term.end_date)
        ).aggregate(total=Sum('payments__amount_paying'))['total'] or Decimal('0.00')

    # Calculate precise collection efficiency percentages securely
    collection_percentage = 0.0
    if total_due_sum > 0:
        collection_percentage = round((float(collected_sum) / float(total_due_sum)) * 100, 2)

    # ==========================================
    # SECTION 2: PROCUREMENT AND EXPENSES AGGREGATION
    # ==========================================
    requisitions_query = Requisition.global_objects.filter(campus=campus, term=term)
    
    req_metrics = requisitions_query.aggregate(
        total_count=Count('id'),
        approved_count=Count('id', filter=Q(status='APPROVED'))
    )

    line_items = RequisitionItem.objects.filter(requisition__campus=campus, requisition__term=term)
    line_cost_expr = ExpressionWrapper(F('price_per_unit') * F('quantity'), output_field=DecimalField())
    
    total_requested = line_items.aggregate(total=Sum(line_cost_expr))['total'] or Decimal('0.00')
    total_funded = line_items.filter(requisition__status='APPROVED').aggregate(
        total=Sum(line_cost_expr)
    )['total'] or Decimal('0.00')

    category_breakdown = line_items.filter(requisition__status='APPROVED').values(
        'category__title'
    ).annotate(
        total_spent=Sum(line_cost_expr)
    ).order_by('-total_spent')

    return {
        'campus': campus, 'term': term,
        'active_pupils': active_count, 'on_transport': transport_count,
        'scholarship_pupils_count': scholarship_pupils_count,
        'total_scholarships_given': scholarship_amt_sum,
        'total_arrears': arrears_sum, 
        'total_base_fees': base_term_fees,
        'total_revenue_transport': transport_sum, 
        'total_revenue_extras': extras_sum, 
        'total_fees_billed': pure_term_billings,
        'total_due_grand': total_due_sum, 
        'total_collected': collected_sum,
        'percentage_collection': collection_percentage,
        'num_requisitions': req_metrics['total_count'] or 0,
        'num_requisitions_approved' : req_metrics['approved_count'] or 0,
        'total_requested_amt': total_requested, 
        'total_approved_funded_amt': total_funded,
        'category_breakdown': category_breakdown
    }



@login_required
def term_summary_dashboard(request):
    """👑 Web Dashboard View: Financial Evaluation Reports Grid Entry Panel"""
    
    if not request.user.profile.is_director and not request.user.profile.is_global_admin:
        messages.error(request, "Access Denied: Analytics dashboards are restricted to Directors and Founders.")
        return redirect('expense_report')

    # 1. Capture incoming targets from your Growth Analytics links
    campus_id = request.GET.get('campus') or getattr(request.user.profile.campus, 'id', None) or Campus.objects.first().id
    term_id = request.GET.get('term') or Term.objects.filter(is_current=True).first().id

    # 2. Extract calculations matrix data (This now inherits the scholarship stats automatically)
    context = get_term_summary_data(campus_id, term_id)
    
    # 3. Pull all available campuses
    context['all_campuses'] = Campus.objects.all()
    
    # 4. Query using .objects but restrict to terms with active invoice footprints
    historical_term_ids = Invoice.objects.values_list('term_id', flat=True).distinct()
    context['all_terms'] = Term.objects.filter(id__in=historical_term_ids) 

    return render(request, 'students/term_summary_report.html', context)



@login_required
def export_term_summary_excel(request):
    """👑 OpenPyXL Excel Pipeline: Generates the specialized multi-section matrix report"""
    
    if not request.user.profile.is_director and not request.user.profile.is_global_admin:
        messages.error(request, "Access Denied: Analytics dashboards are restricted to Directors and Founders.")
        return redirect('revenue_dashboard')

    campus_id = request.GET.get('campus')
    term_id = request.GET.get('term')
    
    data = get_term_summary_data(campus_id, term_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Term Summary {data['term']} " 
    ws.views.sheetView[0].showGridLines = True

    # High-End Corporate Theme Styling Constants
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    dark_purple_fill = PatternFill(start_color="4A148C", end_color="4A148C", fill_type="solid")
    accent_bar_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    
    font_title = Font(name="Calibri", size=16, bold=True, color="1E3A8A")
    font_section = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    font_bold = Font(name="Calibri", size=11, bold=True)
    font_normal = Font(name="Calibri", size=11)
    
    thin_border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'),
                         top=Side(style='thin', color='CBD5E1'), bottom=Side(style='thin', color='CBD5E1'))

    # Title Meta Block Header
    ws['A1'] = f"Skylon Management Systems — Operational Status Report"
    ws['A1'].font = font_title
    ws['A2'] = f"Scope: {data['campus'].name} | Academic Timeline Frame: {data['term']}"
    ws['A2'].font = Font(name="Calibri", size=11, italic=True, color="475569")

    # ==========================================
    # SECTION 1 WRITE: REVENUE AND LEDGER TRACKING
    # ==========================================
    ws.append([]) # Blank spacer row
    ws.append(["SECTION A: REVENUE & LEDGER COLLECTION BALANCES"])
    ws.merge_cells('A4:B4')
    ws.cell(row=4, column=1).fill = navy_fill; ws.cell(row=4, column=1).font = font_section

    rev_rows = [
        ("Total Active Registry Pupils", data['active_pupils'], "#,##0"),
        ("Active Pupils Logged on Transport", data['on_transport'], "#,##0"),
        ("Number of Pupils Awarded Scholarship", data['scholarship_pupils_count'], "#,##0"),

        ("Pure Base Tuition Fees Generated", float(data['total_base_fees']), "#,##0.00"),
        ("Previous Outstanding Arrears Carried Forward", float(data['total_arrears']), "#,##0.00"),
        ("Total Billed Transport Fees Revenue", float(data['total_revenue_transport']), "#,##0.00"),
        ("Total Billed Extras / Mandatory Items Revenue", float(data['total_revenue_extras']), "#,##0.00"),
        ("Total Awarded Sponsorship/Scholarship", float(data['total_scholarships_given']), "#,##0.00"),

        ("Grand Total Expected Financial Obligation", float(data['total_due_grand']), "#,##0.00"),
        ("Total Revenue Collections Realized To Date", float(data['total_collected']), "#,##0.00"),
        ("True Collection Efficiency Ratio", float(data['percentage_collection']) / 100.0, "0.00%")
    ]

    current_row = 5
    for label, val, num_format in rev_rows:
        ws.append([label, val])
        ws.cell(row=current_row, column=1).font = font_normal; ws.cell(row=current_row, column=1).border = thin_border
        cell_val = ws.cell(row=current_row, column=2)
        cell_val.font = font_bold; cell_val.border = thin_border; cell_val.number_format = num_format
        cell_val.alignment = Alignment(horizontal="right")
        current_row += 1

    # ==========================================
    # SECTION 2 WRITE: EXPENDITURES & CATEGORIES
    # ==========================================
    current_row += 1 # Added spacing rows layout
    ws.cell(row=current_row, column=1, value="SECTION B: PROCUREMENT EXPENSES & FUNDING AUDITING").font = font_section
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=2)
    ws.cell(row=current_row, column=1).fill = dark_purple_fill
    current_row += 1

    exp_rows = [
        ("Number of Requisitions Filed", data['num_requisitions'], "#,##0"),
        ("Total Estimated Requested Gross Capital", float(data['total_requested_amt']), "#,##0.00"),
        ("Number of Requisitions Approved", data['num_requisitions_approved'], "#,##0"),
        ("Total Approved & Funded Disbursements Allocations", float(data['total_approved_funded_amt']), "#,##0.00"),
    ]

    for label, val, num_format in exp_rows:
        ws.cell(row=current_row, column=1, value=label).font = font_normal
        ws.cell(row=current_row, column=1).border = thin_border
        cell_val = ws.cell(row=current_row, column=2, value=val)
        cell_val.font = font_bold; cell_val.border = thin_border; cell_val.number_format = num_format
        cell_val.alignment = Alignment(horizontal="right")
        current_row += 1

    # Category Breakdown Header Row Layout Integration
    current_row += 1
    ws.cell(row=current_row, column=1, value="Funded Expense Category").font = font_bold
    ws.cell(row=current_row, column=1).fill = accent_bar_fill; ws.cell(row=current_row, column=1).border = thin_border
    ws.cell(row=current_row, column=2, value="Disbursed Sum").font = font_bold
    ws.cell(row=current_row, column=2).fill = accent_bar_fill; ws.cell(row=current_row, column=2).border = thin_border
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="right")
    current_row += 1

    for cat in data['category_breakdown']:
        ws.cell(row=current_row, column=1, value=cat['category__title']).font = font_normal
        ws.cell(row=current_row, column=1).border = thin_border
        cell_spent = ws.cell(row=current_row, column=2, value=float(cat['total_spent'] or 0))
        cell_spent.font = font_normal; cell_spent.border = thin_border; cell_spent.number_format = "#,##0.00"
        cell_spent.alignment = Alignment(horizontal="right")
        current_row += 1

    # Scale wide cells safely
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 24

    current_date = timezone.now().strftime('%Y-%m-%d')
    filename = f"{data['term']}_Executive_Summary_{data['campus'].name.replace(' ', '_')}_{current_date}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




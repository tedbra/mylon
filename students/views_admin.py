# students/admin_views.py
from decimal import Decimal
from django.db import transaction
from django.db.models import Sum
from django.contrib import messages
from django.shortcuts import redirect, render, get_object_or_404
from django.http import HttpResponse, HttpResponseForbidden
from django.contrib.auth.decorators import login_required
from .models import Term, Student, Invoice, GradeTermFee, PaymentHistory,Campus

#For exporting data to Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils import timezone # 🎯 Added for dynamic filename dating


@login_required
def execute_automatic_next_term_rollover(request):
    """Production-grade rollover tool. Processes timeline states entirely in the background."""
    # 🔒 Executive Guardrail
    user_profile = getattr(request.user, 'profile', None)
    if user_profile and not user_profile.is_global_admin:
        messages.error(request, "Access Denied.")
        return redirect('revenue_dashboard')

    if request.method == "POST":
        with transaction.atomic():
            old_term = Term.objects.filter(is_current=True).first()
            if not old_term:
                messages.error(request, "Rollover failed: No current active term found to roll over from.")
                return redirect('revenue_dashboard')

            next_term = Term.objects.filter(start_date__gt=old_term.start_date).order_by('start_date').first()
            if not next_term:
                messages.error(request, f"Rollover aborted: No future term has been created in the calendar after {old_term}.")
                return redirect('revenue_dashboard')

            # Advance state flags
            old_term.is_current = False
            old_term.save()
            next_term.is_current = True
            next_term.save()

            if next_term.session != old_term.session:
                old_term.session.is_current = False
                old_term.session.save()
                next_term.session.is_current = True
                next_term.session.save()

            active_students = Student.objects.filter(status='ACTIVE')
            invoices_created = 0
            
            for student in active_students:
                previous_arrears = Decimal('0.00')
                old_invoice = Invoice.objects.filter(student=student, term=old_term).first()
                student.scholarship_status = "NONE"
                student.scholarship_amount = Decimal('0:00')
                student.proposed_scholarship_amount = Decimal('0:00')
                student.save()
                
                if old_invoice:
                    total_paid_old_term = PaymentHistory.global_objects.filter(
                        student=student,
                        date_paid__range=(old_term.start_date, old_term.end_date)
                    ).aggregate(total=Sum('amount_paying'))['total'] or Decimal('0.00')
                    
                    outstanding_balance = old_invoice.amount - total_paid_old_term
                    if outstanding_balance > 0:
                        previous_arrears = outstanding_balance



                fee_rule = GradeTermFee.objects.filter(grade=student.grade, term=next_term).first()
                tuition_amount = Decimal(str(fee_rule.amount)) if fee_rule else Decimal('0.00')
                
                if not Invoice.objects.filter(student=student, term=next_term).exists():
                    Invoice.objects.create(
                        student=student,
                        term=next_term,
                        previous_arrears=previous_arrears,
                        tuition_amount=Decimal('0.00'),   # Temporary placeholders
                        transport_amount=Decimal('0.00'), # Temporary placeholders
                        extras_amount=Decimal('0.00'),    # Temporary placeholders
                        scholarship_applied=Decimal('0.00'),
                        amount=previous_arrears           # Initial total is just the debt
                    )

                    student.sync_current_term_invoice()
                    invoices_created += 1

            messages.success(request, f"Timeline advanced! {old_term} closed out. {next_term} is now active. Billed {invoices_created} students.")
            return redirect('revenue_dashboard')

    # Fallback guard against accidental GET clicks
    return redirect('revenue_dashboard')


@login_required
def recalculate_current_term_fees(request):
    """Safely corrects tuition fees for the current active term mid-stream and updates all itemized fields."""
    user_profile = getattr(request.user, 'profile', None)
    if user_profile and not user_profile.is_global_admin:
        messages.error(request, "Access Denied.")
        return redirect('revenue_dashboard')

    if request.method == "POST":
        current_term = Term.objects.filter(is_current=True).first()
        if not current_term:
            messages.error(request, "Update failed: No active term found.")
            return redirect('revenue_dashboard')

        # 1. Target invoices for the current active period
        current_invoices = Invoice.objects.filter(term=current_term).select_related('student')
        updates_count = 0

        with transaction.atomic():
            if not current_invoices:
                old_term = Term.objects.filter(end_date__lt=current_term.start_date).order_by('-end_date').first()
                if not old_term:
                    active_students = Student.objects.filter(status='ACTIVE')
                    invoices_created = 0
                    
                    for student in active_students:
                        previous_arrears = Decimal('0.00')
                        fee_rule = GradeTermFee.objects.filter(grade=student.grade, term=current_term).first()
                        tuition_amount = Decimal(str(fee_rule.amount)) if fee_rule else Decimal('0.00')
                        
                        if not Invoice.objects.filter(student=student, term=current_term).exists():
                            Invoice.objects.create(
                                student=student,
                                term=current_term,
                                previous_arrears=previous_arrears,
                                tuition_amount=Decimal('0.00'),   # Temporary placeholders
                                transport_amount=Decimal('0.00'), # Temporary placeholders
                                extras_amount=Decimal('0.00'),    # Temporary placeholders
                                scholarship_applied=Decimal('0.00'), 
                                amount=previous_arrears           # Initial total is just the debt
                            )

                            student.sync_current_term_invoice()
                            invoices_created += 1

                    messages.success(request, f"No Previous Invoices, No Current Invoice - Created {invoices_created} New Invoices in {current_term}.")
                    return redirect('revenue_dashboard')
                else:
                    # Advance state flags
                    active_students = Student.objects.filter(status='ACTIVE')
                    invoices_created = 0
                    
                    for student in active_students:
                        previous_arrears = Decimal('0.00')
                        old_invoice = Invoice.objects.filter(student=student, term=old_term).first()
                        
                        if old_invoice:
                            total_paid_old_term = PaymentHistory.global_objects.filter(
                                student=student,
                                date_paid__range=(old_term.start_date, old_term.end_date)
                            ).aggregate(total=Sum('amount_paying'))['total'] or Decimal('0.00')
                            
                            outstanding_balance = old_invoice.amount - total_paid_old_term
                            if outstanding_balance > 0:
                                previous_arrears = outstanding_balance


                        fee_rule = GradeTermFee.objects.filter(grade=student.grade, term=current_term).first()
                        tuition_amount = Decimal(str(fee_rule.amount)) if fee_rule else Decimal('0.00')
                        
                        if not Invoice.objects.filter(student=student, term=current_term).exists():
                            Invoice.objects.create(
                                student=student,
                                term=current_term,
                                previous_arrears=previous_arrears,
                                tuition_amount=Decimal('0.00'),   # Temporary placeholders
                                transport_amount=Decimal('0.00'), # Temporary placeholders
                                scholarship_applied=Decimal('0.00'),
                                extras_amount=Decimal('0.00'),    # Temporary placeholders
                                amount=previous_arrears           # Initial total is just the debt
                            )

                            student.sync_current_term_invoice()
                            invoices_created += 1
                    messages.success(request, f"Previous Invoices, But no Current Invoice - Created {invoices_created} New Invoices in {current_term}.")
                    return redirect('revenue_dashboard')


            for invoice in current_invoices:
                # 2. 🔥 Leverage the exact same sync method!
                # This automatically fetches the updated GradeTermFee, retains transport,
                # retains extras, preserves the frozen arrears snapshot, and recalculates everything.
                invoice.student.sync_current_term_invoice()
                updates_count += 1

        messages.success(request, f"Successfully recalculated and itemized profiles for {updates_count} statements in {current_term}.")
        return redirect('revenue_dashboard')

    return redirect('revenue_dashboard')


@login_required
def admin_work(request):
    """Admin interface for executing term rollovers and fee recalculations."""
    user_profile = getattr(request.user, 'profile', None)
    if user_profile and not user_profile.is_global_admin:
        messages.error(request, "Access Denied.")
        return redirect('revenue_dashboard')

    context = {
        'is_global_admin': user_profile.is_global_admin if user_profile else False,
    }
    return render(request, 'students/admin_work.html', context)


@login_required
def export_students_excel(request):
    """
    Generates an Excel workbook with sheets per campus, optionally 
    filtering rows based on client-side UI visibility states.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    user_profile = request.user.profile
    if user_profile.is_director or user_profile.is_global_admin:
        campuses = Campus.objects.all()
    else:
        campuses = Campus.objects.filter(id=user_profile.campus.id) if user_profile.campus else Campus.objects.none()

    # 🎯 NEW: Read the list of student IDs passed from JavaScript
    student_ids_param = request.GET.get('student_ids', None)
    visible_student_ids = student_ids_param.split(',') if student_ids_param else None

    # Common styling objects
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    normal_font = Font(name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
                         top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    headers = ["ID", "Full Name", "Campus Branch", "Grade Allocation", "Total Owing", "Paid", "Remaining Balance", "Contact", "Transport", "Status", "Media Rights"]

    for campus in campuses:
        sheet_title = "".join(c for c in campus.name if c.isalnum() or c in "._- ")[:31].strip() or f"Campus {campus.id}"
        
        # Base filter query for this campus tab
        student_query = Student.objects.filter(campus=campus).select_related('grade')
        
        # 🎯 NEW: Apply the ID subset filter if the request originated from the filtered UI list
        if visible_student_ids is not None:
            student_query = student_query.filter(student_id__in=visible_student_ids)

        # Skip creating a campus sheet entirely if it contains no filtered students
        if not student_query.exists():
            continue

        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = white_bold; cell.fill = navy_fill; cell.border = thin_border

        for row_idx, student in enumerate(student_query, 2):
            row_data = [
                student.student_id, student.name, campus.name, student.grade.title if student.grade else "N/A",
                float(student.current_term_total_due or 0), float(student.current_term_paid or 0), float(student.current_term_outstanding_balance or 0),
                student.phone_number_1, "Uses Transport" if student.needs_transport else "No Transport", student.status, "Allowed" if student.image_use_consent else "Restricted"
            ]
            ws.append(row_data)

            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = normal_font; cell.border = thin_border
                if is_even: cell.fill = zebra_fill
                if col_idx in [5, 6, 7]:
                    cell.number_format = '#,##0.00'; cell.alignment = Alignment(horizontal="right")

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Matches Found")
        ws.append(["No records matched your specific filter selections."])

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="Filtered_Student_Registry.xlsx"'
    wb.save(response)
    return response


@login_required
def xexport_students_excel_all(request):
    """
    Generates an exhaustive Master Excel archive file mapping exact 
    fields from the Student model schema, categorized into campus tabs.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Resolve multi-campus permission scopes
    user_profile = request.user.profile
    if user_profile.is_director or user_profile.is_global_admin:
        campuses = Campus.objects.all()
    else:
        campuses = Campus.objects.filter(id=user_profile.campus.id) if user_profile.campus else Campus.objects.none()

    # Capture client-side list view subsets if applicable
    student_ids_param = request.GET.get('student_ids', None)
    visible_student_ids = student_ids_param.split(',') if student_ids_param else None

    # 2. Design Schemes
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    normal_font = Font(name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
                         top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    # Exact columns mapping to your model attributes
    headers = [
        "Student ID", "Full Name", "Campus Branch", "Grade Level", "Gender", "Status", "Date of Birth", "Joined Intake",
        "Carried Arrears", "Termly Fees Billed", "Transport Fee", "Grand Total Due", "Total Paid to Date", "Outstanding Balance",
        "Transport Setup", "Transport Route", "Media Consent", "Parent/Guardian", "Primary Phone", "Secondary Phone", "Address Setup"
    ]

    # Indices matching financial columns (1-indexed for Excel loops)
    financial_column_indices = [9, 10, 11, 12, 13, 14]

    # 3. Process Campuses
    for campus in campuses:
        sheet_title = "".join(c for c in campus.name if c.isalnum() or c in "._- ")[:31].strip() or f"Campus {campus.id}"
        
        # Pull records cleanly matching isolation constraints
        student_query = Student.objects.filter(campus=campus).select_related('grade', 'transport_route')

        if visible_student_ids is not None:
            student_query = student_query.filter(student_id__in=visible_student_ids)

        if not student_query.exists():
            continue

        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = white_bold; cell.fill = navy_fill; cell.border = thin_border

        # 4. Process Records
        for row_idx, student in enumerate(student_query, 2):
            
            # Format dates nicely
            dob_str = student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else "N/A"
            intake_str = f"{student.enrollment_year} - Term {student.enrollment_term}" if student.enrollment_year else "N/A"
            address_block = f"{student.address}, {student.city}, {student.country}"

            row_data = [
                student.student_id,
                student.name,
                campus.name,
                student.grade.title if student.grade else "N/A",
                student.get_gender_display(),
                student.get_status_display(),
                dob_str,
                intake_str,
                
                # Dynamic calculated ledger property outputs mapped explicitly from your models
                float(student.current_term_arrears or 0),
                float(student.current_term_fees_billed or 0),
                float(student.current_term_tranport or 0),
                float(student.current_term_total_due or 0),
                float(student.current_term_paid or 0),
                float(student.current_term_outstanding_balance or 0),
                
                "Required" if student.needs_transport else "No",
                student.transport_route.route_name if student.transport_route else "N/A",
                "Allowed" if student.image_use_consent else "Restricted",
                student.parent_name,
                student.phone_number_1,
                student.phone_number_2 or "N/A",
                address_block
            ]
            ws.append(row_data)

            # Apply Styles & Formats per row cell
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = normal_font; cell.border = thin_border
                if is_even: 
                    cell.fill = zebra_fill

                if col_idx in financial_column_indices:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        # Column scaling adjustments loop execution
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Records Available")
        ws.append(["No records matching your current active filter queries found."])

    # 5. Build dynamically dated attachment name payload
    current_date = timezone.now().strftime('%Y-%m-%d')
    filename = f"Skylon_Master_Archive_{current_date}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_students_excel_all(request):
    """
    Generates an exhaustive Master Excel archive file mapping exact 
    fields from the Student model schema, categorized into campus tabs.
    Includes active scholarship metadata and updated ledger financial tracks.
    """
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Resolve multi-campus permission scopes
    user_profile = request.user.profile
    if user_profile.is_director or user_profile.is_global_admin:
        campuses = Campus.objects.all()
    else:
        campuses = Campus.objects.filter(id=user_profile.campus.id) if user_profile.campus else Campus.objects.none()

    # Capture client-side list view subsets if applicable
    student_ids_param = request.GET.get('student_ids', None)
    visible_student_ids = student_ids_param.split(',') if student_ids_param else None

    # 2. Design Schemes
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    white_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    normal_font = Font(name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin', color='E2E8F0'), right=Side(style='thin', color='E2E8F0'),
                         top=Side(style='thin', color='E2E8F0'), bottom=Side(style='thin', color='E2E8F0'))

    # Exact columns mapping to your model attributes (Scholarship tracking inserted sequentially)
    headers = [
        "Student ID", "Full Name", "Campus Branch", "Grade Level", "Gender", "Status", "Date of Birth", "Joined Intake",
        "Scholarship Status", "Scholarship Allocated", "Carried Arrears", "Termly Fees Billed", "Transport Fee", 
        "Grand Total Due", "Total Paid to Date", "Outstanding Balance", "Transport Setup", "Transport Route", 
        "Media Consent", "Parent/Guardian", "Primary Phone", "Secondary Phone", "Address Setup"
    ]

    # Indices matching financial columns (1-indexed for Excel loops)
    # Updated to reflect the shifted indices due to newly added tracking blocks
    financial_column_indices = [10, 11, 12, 13, 14, 15, 16]

    # 3. Process Campuses
    for campus in campuses:
        sheet_title = "".join(c for c in campus.name if c.isalnum() or c in "._- ")[:31].strip() or f"Campus {campus.id}"
        
        # Pull records cleanly matching isolation constraints
        student_query = Student.objects.filter(campus=campus).select_related('grade', 'transport_route')

        if visible_student_ids is not None:
            student_query = student_query.filter(student_id__in=visible_student_ids)

        if not student_query.exists():
            continue

        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        ws.append(headers)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = white_bold; cell.fill = navy_fill; cell.border = thin_border

        # 4. Process Records
        for row_idx, student in enumerate(student_query, 2):
            
            # Format dates nicely
            dob_str = student.date_of_birth.strftime('%Y-%m-%d') if student.date_of_birth else "N/A"
            intake_str = f"{student.enrollment_year} - Term {student.enrollment_term}" if student.enrollment_year else "N/A"
            address_block = f"{student.address}, {student.city}, {student.country}"

            row_data = [
                student.student_id,
                student.name,
                campus.name,
                student.grade.title if student.grade else "N/A",
                student.get_gender_display(),
                student.get_status_display(),
                dob_str,
                intake_str,
                
                # 🎓 Scholarship status tracking mapping blocks
                student.get_scholarship_status_display() if student.scholarship_status else "None",
                float(student.scholarship_amount or 0),
                
                # Dynamic calculated ledger property outputs mapped explicitly from your models
                float(student.current_term_arrears or 0),
                float(student.current_term_fees_billed or 0),
                float(student.current_term_tranport or 0),
                float(student.current_term_total_due or 0),
                float(student.current_term_paid or 0),
                float(student.current_term_outstanding_balance or 0),
                
                "Required" if student.needs_transport else "No",
                student.transport_route.route_name if student.transport_route else "N/A",
                "Allowed" if student.image_use_consent else "Restricted",
                student.parent_name,
                student.phone_number_1,
                student.phone_number_2 or "N/A",
                address_block
            ]
            ws.append(row_data)

            # Apply Styles & Formats per row cell
            is_even = (row_idx % 2 == 0)
            for col_idx in range(1, len(row_data) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = normal_font; cell.border = thin_border
                if is_even: 
                    cell.fill = zebra_fill

                if col_idx in financial_column_indices:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal="right")

        # Column scaling adjustments loop execution
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    if len(wb.sheetnames) == 0:
        ws = wb.create_sheet(title="No Records Available")
        ws.append(["No records matching your current active filter queries found."])

    # 5. Build dynamically dated attachment name payload
    current_date = timezone.now().strftime('%Y-%m-%d')
    filename = f"Skylon_Master_Archive_{current_date}.xlsx"

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def review_scholarship(request, student_id, action):
    """👑 Administrative Approval Engine with Traceability Audit Logging"""
    user_profile = getattr(request.user, 'profile', None)
    if not getattr(user_profile, 'is_global_admin', False) and not getattr(user_profile, 'is_director', False):
        return HttpResponseForbidden("Access Denied: Administrative credentials required.")

    student = get_object_or_404(Student, id=student_id)

    if action == 'approve':
        student.scholarship_status = 'APPROVED'
        student.scholarship_amount = student.proposed_scholarship_amount
        student.approved_by = request.user  # 🎯 Stamp the specific user ID here
        messages.success(request, f"Scholarship approved for {student.first_name}.")
        
    elif action == 'reject':
        student.scholarship_status = 'REJECTED'
        student.scholarship_amount = Decimal('0.00')
        student.proposed_scholarship_amount = Decimal('0.00')
        student.approved_by = request.user  # 🎯 Even rejections get logged for accountability
        messages.warning(request, f"Scholarship request denied for {student.first_name}.")

    elif action == 'revoke':
        student.scholarship_status = 'NONE'
        student.scholarship_amount = Decimal('0.00')
        student.proposed_scholarship_amount = Decimal('0.00')
        student.approved_by = request.user  # 🎯 Log the admin who revoked it
        messages.warning(request, f"Scholarship successfully revoked for {student.first_name}.")

    student.save()
    return redirect('global_analytics_dashboard')



@login_required
def scholarship_review_dashboard(request):
    """👑 Global Admin Dashboard to view, approve, reject, or revoke scholarships."""
    user_profile = getattr(request.user, 'profile', None)
    if not user_profile or not (user_profile.is_director or getattr(user_profile, 'is_global_admin', False)):
        return HttpResponseForbidden("Access Denied: Administrative credentials required.")

    if request.method == "POST":
        student_id = request.POST.get("student_id")
        action = request.POST.get("action")
        student = get_object_or_404(Student, id=student_id)

        if action == "approve":
            student.scholarship_status = 'APPROVED'
            student.scholarship_amount = student.proposed_scholarship_amount
            student.approved_by = request.user
            messages.success(request, f"Successfully approved scholarship for {student.name}.")

        elif action == "reject":
            student.scholarship_status = 'REJECTED'
            student.scholarship_amount = Decimal('0.00')
            student.proposed_scholarship_amount = Decimal('0.00')
            student.approved_by = request.user
            messages.warning(request, f"Rejected scholarship application for {student.name}.")

        # 🎯 NEW ACTION: REVOKE ACTIVE SCHOLARSHIP
        elif action == "revoke":
            student.scholarship_status = 'NONE'
            student.scholarship_amount = Decimal('0.00')
            student.proposed_scholarship_amount = Decimal('0.00')
            student.approved_by = request.user  # Trailed log of who revoked it
            messages.error(request, f"Scholarship access has been permanently revoked for {student.name}.")

        student.save()
        
        # Instantly updates or rolls back their current term billing invoice statement rows
        student.sync_current_term_invoice()
        return redirect('scholarship_review_dashboard')

    # Read View: Exclude 'NONE' to show all active lifecycle items (PENDING, APPROVED, REJECTED)
    all_scholarships = Student.global_objects.exclude(scholarship_status='NONE').select_related('campus', 'grade')
    
    context = {
        'scholarships': all_scholarships,
    }
    return render(request, 'students/scholarship_review.html', context)


